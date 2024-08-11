class ExcelInterface:
    def __init__(self, file_path):
        import openpyxl
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def find_keyword_value(self, keyword):
        for row in self.sheet.iter_rows():
            for cell in row:
                if str(cell.value).lower() == str(keyword).lower():
                    return cell.offset(column=1).value
        return None

    def find_keyword_all_right(self, keyword):
        results = []
        for row in self.sheet.iter_rows():
            for cell in row:
                if str(cell.value).lower() == keyword.lower():
                    col = cell.column + 1
                    while col <= self.sheet.max_column:
                        right_cell = self.sheet.cell(row=cell.row, column=col)
                        if right_cell.value is not None:
                            results.append(right_cell.value)
                        col += 1
                    return results
        return None

    def rows_to_dict(self, rows):
        result_dict = {}
        for row_num in rows:
            row_data = []
            key = None
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row_num, column=col).value
                if key is None and cell_value is not None:
                    key = cell_value
                elif key is not None and cell_value is not None:
                    row_data.append(cell_value)
            if key is not None:
                result_dict[key] = row_data
        return result_dict

    def get_rows_between_keyword_and_empty(self, keyword):
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == keyword:
                    start_row = cell.row + 1
                    end_row = start_row
                    while end_row <= self.sheet.max_row and self.sheet.cell(row=end_row, column=cell.column).value is not None:
                        end_row += 1
                    return (start_row, end_row - 1)
        return None

if __name__ == '__main__':
    import numpy as np
    file_path = 'ArbitraryTestbench.xlsx'
    excel_interface = ExcelInterface(file_path)
    clk_period = excel_interface.find_keyword_value("clk_period")
    print(f"clk_period: {clk_period}")

    SA_CLK_VALUES = excel_interface.find_keyword_all_right("SA_CLK")
    print(f"SA_CLK_VALUES: {SA_CLK_VALUES}")

    get_rows = excel_interface.get_rows_between_keyword_and_empty("Arbitrary Waveforms")
    print(f"get_rows: {get_rows}")
    row_dict = excel_interface.rows_to_dict(np.arange(get_rows[0], get_rows[1] + 1))
    print(row_dict)
    quit()
# Example usage:
# file_path = "your_excel_file.xlsx"
# excel_interface = ExcelInterface(file_path)
# value = excel_interface.find_keyword_value("clk_period")
# print(value)
# values_right = excel_interface.find_keyword_all_right("clk_period")
# print(values_right)
# rows_dict = excel_interface.rows_to_dict([1, 2, 3])
# print(rows_dict)
